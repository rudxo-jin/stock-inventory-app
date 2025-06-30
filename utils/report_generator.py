import pandas as pd
import numpy as np
from datetime import datetime
import os
from typing import Dict, Tuple, Optional

class ReportGenerator:
    """재고조사 보고서 생성 클래스"""
    
    def __init__(self):
        self.report_data = None
        self.summary_stats = None
        # 다중 시트를 위한 데이터 저장
        self.part_data = None
        self.inventory_data = None
        self.final_data = None
        self.adjustment_data = None
    
    def generate_report_data(
        self, 
        inventory_data: pd.DataFrame,
        store_info: Dict,
        part_data: Optional[pd.DataFrame] = None,
        final_data: Optional[pd.DataFrame] = None,
        adjustment_summary: Optional[Dict] = None
    ) -> Dict:
        """
        보고서 데이터 생성
        
        Args:
            inventory_data: 실재고 처리 데이터 (필수)
            store_info: 점포 정보 (필수)
            part_data: 원본 PART 데이터 (선택)
            final_data: 최종 재고 데이터 (재고조정 적용 후, 선택)
            adjustment_summary: 재고조정 요약 (선택)
            
        Returns:
            보고서 데이터 딕셔너리
        """
        
        # final_data가 없으면 inventory_data 사용
        if final_data is None:
            final_data = inventory_data
            
        # part_data가 없으면 inventory_data에서 추정
        if part_data is None:
            # inventory_data에서 전산재고 정보 추출
            part_data = inventory_data.copy()
        
        # adjustment_summary가 없으면 기본값 사용
        if adjustment_summary is None:
            adjustment_summary = {
                'positive_amount': 0,
                'negative_amount': 0,
                'total_adjustments': 0
            }
        
        # 다중 시트를 위한 데이터 저장
        self.part_data = part_data
        self.inventory_data = inventory_data
        self.final_data = final_data
        
        # 1. 전산재고 vs 실재고 비교
        inventory_comparison = self._calculate_inventory_comparison(part_data, inventory_data)
        
        # 2. 재고조정 영향 계산
        adjustment_impact = self._calculate_adjustment_impact(inventory_data, final_data, adjustment_summary)
        
        # 3. 총 재고차액 계산
        total_impact = self._calculate_total_impact(inventory_comparison, adjustment_impact)
        
        # 4. 보고서 데이터 구성
        report_data = {
            'store_info': store_info,
            'inventory_comparison': inventory_comparison,
            'adjustment_impact': adjustment_impact,
            'total_impact': total_impact,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        self.report_data = report_data
        return report_data
    
    def set_adjustment_data(self, adjustment_data: pd.DataFrame):
        """재고조정 데이터 설정 (다중 시트용)"""
        self.adjustment_data = adjustment_data
        
        # 웹 배포용: 디버그 정보 제거
    
    def _calculate_inventory_comparison(self, part_data: pd.DataFrame, inventory_data: pd.DataFrame) -> Dict:
        """전산재고 vs 실재고 비교 계산"""
        
        # 전산재고액 계산
        if '재고액' in part_data.columns:
            computer_stock_value = part_data['재고액'].sum()
        else:
            # inventory_data에서 전산재고액 추정 (재고 * 단가)
            computer_stock_value = (inventory_data['재고'] * inventory_data['단가']).sum()
        
        # 실재고 증가/감소 분리
        positive_diff = inventory_data[inventory_data['차이'] > 0]
        negative_diff = inventory_data[inventory_data['차이'] < 0]
        
        positive_amount = positive_diff['차액'].sum() if not positive_diff.empty else 0
        negative_amount = abs(negative_diff['차액'].sum()) if not negative_diff.empty else 0  # 절댓값
        
        # 최종재고액
        final_stock_value = computer_stock_value + positive_amount - negative_amount
        
        # 차액 (실재고 - 전산재고)
        difference = positive_amount - negative_amount
        
        return {
            'computer_stock_value': computer_stock_value,
            'positive_amount': positive_amount,
            'negative_amount': negative_amount,
            'final_stock_value': final_stock_value,
            'difference': difference
        }
    
    def _calculate_adjustment_impact(self, inventory_data: pd.DataFrame, final_data: pd.DataFrame, adj_summary: Dict) -> Dict:
        """재고조정 영향 계산"""
        
        # adj_summary가 None이면 기본값 사용
        if adj_summary is None:
            adj_summary = {}
        
        positive_adj = adj_summary.get('positive_amount', 0)
        negative_adj = abs(adj_summary.get('negative_amount', 0))  # 절댓값
        adjustment_diff = positive_adj - negative_adj
        
        return {
            'positive_adjustment': positive_adj,
            'negative_adjustment': negative_adj,
            'adjustment_difference': adjustment_diff
        }
    
    def _calculate_total_impact(self, inventory_comp: Dict, adjustment_imp: Dict) -> Dict:
        """총 재고차액 계산"""
        
        total_positive = inventory_comp['positive_amount'] + adjustment_imp['positive_adjustment']
        total_negative = inventory_comp['negative_amount'] + adjustment_imp['negative_adjustment']
        total_difference = total_positive - total_negative
        
        return {
            'total_positive': total_positive,
            'total_negative': total_negative,
            'total_difference': total_difference
        }
    
    def create_excel_report(self) -> bytes:
        """다중 시트 엑셀 보고서 생성 (메모리에서 바이트로 반환)"""
        
        if self.report_data is None:
            raise ValueError("먼저 generate_report_data()를 실행해주세요.")
        
        # 메모리에서 엑셀 파일 생성
        from io import BytesIO
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # 1. 요약 보고서 시트
            summary_df = self._create_summary_sheet()
            summary_df.to_excel(writer, sheet_name='재고조사요약', index=False, header=False)
            
            # 2. 재고차이리스트 (-) 시트
            if self.inventory_data is not None:
                negative_diff_df = self._create_negative_diff_sheet()
                if not negative_diff_df.empty:
                    negative_diff_df.to_excel(writer, sheet_name='재고차이리스트(-)', index=False)
            
            # 3. 재고차이리스트 (+) 시트
            if self.inventory_data is not None:
                positive_diff_df = self._create_positive_diff_sheet()
                if not positive_diff_df.empty:
                    positive_diff_df.to_excel(writer, sheet_name='재고차이리스트(+)', index=False)
            
            # 4. 재고조정리스트 (+) 시트
            if self.adjustment_data is not None:
                positive_adj_df = self._create_positive_adjustment_sheet()
                if not positive_adj_df.empty:
                    positive_adj_df.to_excel(writer, sheet_name='재고조정리스트(+)', index=False)
            
            # 5. 재고조정리스트 (-) 시트
            if self.adjustment_data is not None:
                negative_adj_df = self._create_negative_adjustment_sheet()
                if not negative_adj_df.empty:
                    negative_adj_df.to_excel(writer, sheet_name='재고조정리스트(-)', index=False)
            
            # 6. PART 원본데이터 시트
            if self.part_data is not None:
                part_df = self._create_part_data_sheet()
                if not part_df.empty:
                    part_df.to_excel(writer, sheet_name='PART원본데이터', index=False)
            
            # 7. 전체재고리스트 시트
            if self.inventory_data is not None:
                full_inventory_df = self._create_full_inventory_list_sheet()
                if not full_inventory_df.empty:
                    full_inventory_df.to_excel(writer, sheet_name='전체재고리스트', index=False)
            
            # 워크시트 스타일링
            workbook = writer.book
            worksheet = writer.sheets['재고조사요약']
            
            # 스타일 임포트
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 색상 정의
            title_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")    # 제목용 진한 네이비
            info_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")   # 연한 파란색
            positive_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # 연한 초록색
            negative_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")  # 연한 빨간색
            total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")    # 연한 노란색
            
            # 폰트 정의
            title_font = Font(name='맑은 고딕', size=14, bold=True, color="FFFFFF")  # 제목용 흰색 글자
            info_font = Font(name='맑은 고딕', size=11, bold=True)
            data_font = Font(name='맑은 고딕', size=10)
            amount_font = Font(name='맑은 고딕', size=10, bold=True)
            
            # 정렬 정의
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            
            # 테두리 정의
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # 컬럼 너비 조정 (2컬럼 구조)
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 20
            
            # 먼저 제목 셀 병합
            worksheet.merge_cells('A1:B1')
            
            # 행별 스타일 적용
            max_row = worksheet.max_row
            
            for row_idx in range(1, max_row + 1):
                cell_a = worksheet[f'A{row_idx}']
                cell_b = worksheet[f'B{row_idx}']
                
                # 제목 행 (1행)
                if row_idx == 1:
                    cell_a.font = title_font
                    cell_a.alignment = center_align
                    cell_a.fill = title_fill
                    cell_a.border = thin_border
                
                # 빈 행 (2행)
                elif row_idx == 2:
                    continue
                
                # 점포 정보 섹션 (3-6행)
                elif 3 <= row_idx <= 6:
                    cell_a.fill = info_fill
                    cell_a.font = info_font
                    cell_a.alignment = left_align
                    cell_a.border = thin_border
                    
                    cell_b.font = data_font
                    cell_b.alignment = left_align
                    cell_b.border = thin_border
                
                # 빈 행 (7행)
                elif row_idx == 7:
                    continue
                
                # 데이터 행들 (8행부터)
                elif cell_a.value and str(cell_a.value).strip():
                    # 금액이 있는 행인지 확인
                    if cell_b.value and str(cell_b.value).replace(',', '').replace('-', '').isdigit():
                        # 양수/음수/합계에 따른 색상 적용
                        if '(+)' in str(cell_a.value):
                            cell_a.fill = positive_fill
                            cell_b.fill = positive_fill
                        elif '(-)' in str(cell_a.value):
                            cell_a.fill = negative_fill
                            cell_b.fill = negative_fill
                        elif '계' in str(cell_a.value) or '차액' in str(cell_a.value):
                            cell_a.fill = total_fill
                            cell_b.fill = total_fill
                        
                        cell_a.font = info_font
                        cell_a.alignment = left_align
                        cell_a.border = thin_border
                        
                        cell_b.font = amount_font
                        cell_b.alignment = right_align
                        cell_b.border = thin_border
                        
                        # 음수 금액은 빨간색으로
                        if str(cell_b.value).startswith('-'):
                            cell_b.font = Font(name='맑은 고딕', size=10, bold=True, color="CC0000")
            
            # 행 높이 조정
            for row in range(1, max_row + 1):
                if row == 1:  # 제목 행은 더 높게
                    worksheet.row_dimensions[row].height = 35
                else:
                    worksheet.row_dimensions[row].height = 25
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_summary_sheet(self) -> pd.DataFrame:
        """요약 보고서 시트 데이터 생성"""
        
        store_info = self.report_data['store_info']
        inv_comp = self.report_data['inventory_comparison']
        adj_imp = self.report_data['adjustment_impact']
        total_imp = self.report_data['total_impact']
        
        # 보고서 구조 생성 (2컬럼 구조로 단순화)
        summary_data = []
        
        # 제목 행
        summary_data.extend([
            ['재고 조사 결과 보고', ''],
            ['', ''],
        ])
        
        # 점포 정보
        summary_data.extend([
            ['점포명:', store_info.get('store_name', '')],
            ['재고조사일시:', store_info.get('survey_date', '')],
            ['조사방식:', store_info.get('survey_method', '')],
            ['조사인원:', store_info.get('survey_staff', '')],
            ['', ''],
        ])
        
        # 전산재고 vs 실재고 섹션
        summary_data.extend([
            [f"{store_info.get('survey_date', '2024년 12월 31일')} 전산 재고액", f"{inv_comp['computer_stock_value']:,.0f}"],
            ['(+) 실재고액', f"{inv_comp['positive_amount']:,.0f}"],
            ['(-) 실재고액', f"-{inv_comp['negative_amount']:,.0f}"],
            ['최종재고액', f"{inv_comp['final_stock_value']:,.0f}"],
            ['차액', f"{inv_comp['difference']:,.0f}"],
            ['', ''],
        ])
        
        # 재고조정 섹션
        summary_data.extend([
            ['(+) 재고조정액', f"{adj_imp['positive_adjustment']:,.0f}"],
            ['(-) 재고조정액', f"-{adj_imp['negative_adjustment']:,.0f}"],
            ['재고조정 차액', f"{adj_imp['adjustment_difference']:,.0f}"],
            ['', ''],
        ])
        
        # 총 재고차액 섹션
        summary_data.extend([
            ['(+) 총재고차액', f"{total_imp['total_positive']:,.0f}"],
            ['(-) 총재고차액', f"-{total_imp['total_negative']:,.0f}"],
            ['총재고차액 계', f"{total_imp['total_difference']:,.0f}"],
        ])
        
        return pd.DataFrame(summary_data, columns=['항목', '금액'])
    
    def _create_negative_diff_sheet(self) -> pd.DataFrame:
        """재고차이리스트 (-) 시트 생성"""
        if self.inventory_data is None:
            return pd.DataFrame()
        
        # 차이가 음수인 데이터만 필터링
        negative_data = self.inventory_data[self.inventory_data['차이'] < 0].copy()
        
        if negative_data.empty:
            return pd.DataFrame()
        
        # 필요한 컬럼만 선택하고 정렬
        result_df = negative_data[[
            '제작사 품번', '부품명', '단가', '재고', '재고액', 
            '실재고', '실재고액', '차액', '차이'
        ]].copy()
        
        # 컬럼명 정리
        result_df.columns = [
            '제작사품번', '부품명', '단가', '재고', '재고액',
            '실재고', '실재고액', '차액', '차이'
        ]
        
        # 차액 기준 오름차순 정렬 (가장 큰 손실부터)
        result_df = result_df.sort_values('차액').reset_index(drop=True)
        
        # 합계 행 추가
        if not result_df.empty:
            total_row = pd.DataFrame({
                '제작사품번': ['합계'],
                '부품명': [''],
                '단가': [''],
                '재고': [result_df['재고'].sum()],
                '재고액': [result_df['재고액'].sum()],
                '실재고': [result_df['실재고'].sum()],
                '실재고액': [result_df['실재고액'].sum()],
                '차액': [result_df['차액'].sum()],
                '차이': [result_df['차이'].sum()]
            })
            result_df = pd.concat([result_df, total_row], ignore_index=True)
        
        return result_df
    
    def _create_positive_diff_sheet(self) -> pd.DataFrame:
        """재고차이리스트 (+) 시트 생성"""
        if self.inventory_data is None:
            return pd.DataFrame()
        
        # 차이가 양수인 데이터만 필터링
        positive_data = self.inventory_data[self.inventory_data['차이'] > 0].copy()
        
        if positive_data.empty:
            return pd.DataFrame()
        
        # 필요한 컬럼만 선택하고 정렬
        result_df = positive_data[[
            '제작사 품번', '부품명', '단가', '재고', '재고액', 
            '실재고', '실재고액', '차액', '차이'
        ]].copy()
        
        # 컬럼명 정리
        result_df.columns = [
            '제작사품번', '부품명', '단가', '재고', '재고액',
            '실재고', '실재고액', '차액', '차이'
        ]
        
        # 차액 기준 내림차순 정렬 (가장 큰 이익부터)
        result_df = result_df.sort_values('차액', ascending=False).reset_index(drop=True)
        
        # 합계 행 추가
        if not result_df.empty:
            total_row = pd.DataFrame({
                '제작사품번': ['합계'],
                '부품명': [''],
                '단가': [''],
                '재고': [result_df['재고'].sum()],
                '재고액': [result_df['재고액'].sum()],
                '실재고': [result_df['실재고'].sum()],
                '실재고액': [result_df['실재고액'].sum()],
                '차액': [result_df['차액'].sum()],
                '차이': [result_df['차이'].sum()]
            })
            result_df = pd.concat([result_df, total_row], ignore_index=True)
        
        return result_df
    
    def _create_positive_adjustment_sheet(self) -> pd.DataFrame:
        """재고조정리스트 (+) 시트 생성"""
        if self.adjustment_data is None:
            return pd.DataFrame()
        
        # 조정구분이 '+'인 데이터만 필터링
        if '조정구분' in self.adjustment_data.columns:
            positive_data = self.adjustment_data[self.adjustment_data['조정구분'] == '+'].copy()
        else:
            # 수량변경에서 +가 포함된 데이터 필터링 (더 안전한 방법)
            mask = (self.adjustment_data['수량변경'].astype(str).str.contains('+', na=False) | 
                   self.adjustment_data['수량변경'].astype(str).str.contains('증가', na=False))
            positive_data = self.adjustment_data[mask].copy()
        
        if positive_data.empty:
            return pd.DataFrame()
        
        result_df = self._process_adjustment_data(positive_data)
        
        # 합계 행 추가
        if not result_df.empty:
            total_row = pd.DataFrame({
                '일자': ['합계'],
                '구분': [''],
                '제작사품번': [''],
                '부품명': [''],
                '수량': [result_df['수량'].sum()],
                '단가': [''],
                '금액': [result_df['금액'].sum()]
            })
            result_df = pd.concat([result_df, total_row], ignore_index=True)
        
        return result_df
    
    def _create_negative_adjustment_sheet(self) -> pd.DataFrame:
        """재고조정리스트 (-) 시트 생성"""
        if self.adjustment_data is None:
            return pd.DataFrame()
        
        # 조정구분이 '-'인 데이터만 필터링
        if '조정구분' in self.adjustment_data.columns:
            negative_data = self.adjustment_data[self.adjustment_data['조정구분'] == '-'].copy()
        else:
            # 수량변경에서 -가 포함된 데이터 필터링 (더 안전한 방법)
            mask = (self.adjustment_data['수량변경'].astype(str).str.contains('-', na=False) | 
                   self.adjustment_data['수량변경'].astype(str).str.contains('감소', na=False))
            negative_data = self.adjustment_data[mask].copy()
        
        if negative_data.empty:
            return pd.DataFrame()
        
        result_df = self._process_adjustment_data(negative_data)
        
        # 합계 행 추가
        if not result_df.empty:
            total_row = pd.DataFrame({
                '일자': ['합계'],
                '구분': [''],
                '제작사품번': [''],
                '부품명': [''],
                '수량': [result_df['수량'].sum()],
                '단가': [''],
                '금액': [result_df['금액'].sum()]
            })
            result_df = pd.concat([result_df, total_row], ignore_index=True)
        
        return result_df
    
    def _process_adjustment_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """재고조정 데이터 공통 처리"""
        result_df = data.copy()
        
        # 필요한 컬럼만 선택
        if '조정구분' in result_df.columns:
            result_df = result_df[['일자', '조정구분', '제작사품번', '부품명', '수량']].copy()
            result_df.columns = ['일자', '구분', '제작사품번', '부품명', '수량']
        else:
            result_df = result_df[['일자', '수량변경', '제작사품번', '부품명', '수량']].copy()
            result_df.columns = ['일자', '구분', '제작사품번', '부품명', '수량']
        
        # 단가와 금액 계산 - dtype 명시적 설정
        result_df = result_df.copy()
        result_df['단가'] = 0.0
        result_df['금액'] = 0.0
        
        # inventory_data나 part_data에서 단가 매칭
        if self.inventory_data is not None:
            for idx, row in result_df.iterrows():
                part_code = row['제작사품번']
                quantity = float(row['수량'])
                
                # inventory_data에서 단가 찾기
                matching_rows = self.inventory_data[self.inventory_data['제작사 품번'] == part_code]
                if not matching_rows.empty:
                    unit_price = float(matching_rows.iloc[0]['단가'])
                    result_df.at[idx, '단가'] = unit_price
                    result_df.at[idx, '금액'] = quantity * unit_price
                elif self.part_data is not None:
                    # part_data에서 단가 찾기
                    matching_parts = self.part_data[self.part_data['제작사 품번'] == part_code]
                    if not matching_parts.empty:
                        # 재고액/재고로 단가 계산
                        stock = float(matching_parts.iloc[0]['재고'])
                        stock_value = float(matching_parts.iloc[0]['재고액'])
                        if stock > 0:
                            unit_price = stock_value / stock
                            result_df.at[idx, '단가'] = unit_price
                            result_df.at[idx, '금액'] = quantity * unit_price
        
        # 일자 기준 정렬
        result_df = result_df.sort_values('일자').reset_index(drop=True)
        
        # 일자 포맷 변경
        result_df['일자'] = pd.to_datetime(result_df['일자']).dt.strftime('%Y-%m-%d')
        
        return result_df
    
    def get_summary_stats(self) -> Dict:
        """요약 통계 반환 (UI 표시용)"""
        
        if self.report_data is None:
            return {}
        
        return {
            'inventory_comparison': self.report_data['inventory_comparison'],
            'adjustment_impact': self.report_data['adjustment_impact'],
            'total_impact': self.report_data['total_impact']
        }
    
    def _create_part_data_sheet(self) -> pd.DataFrame:
        """PART 원본데이터 시트 생성"""
        if self.part_data is None:
            return pd.DataFrame()
        
        # PART 원본 데이터 복사
        result_df = self.part_data.copy()
        
        # 컬럼명 한글로 변경
        column_mapping = {
            '제작사 품번': '제작사품번',
            '부품명': '부품명',
            '재고': '재고',
            '재고액': '재고액',
            '단가': '단가'
        }
        
        # 필요한 컬럼만 선택하고 컬럼명 변경
        if set(column_mapping.keys()).issubset(result_df.columns):
            result_df = result_df[list(column_mapping.keys())].copy()
            result_df.columns = list(column_mapping.values())
        
        # 재고액 기준 내림차순 정렬
        result_df = result_df.sort_values('재고액', ascending=False).reset_index(drop=True)
        
        # 합계 행 추가
        if not result_df.empty:
            total_row = pd.DataFrame({
                '제작사품번': ['합계'],
                '부품명': [''],
                '재고': [result_df['재고'].sum()],
                '재고액': [result_df['재고액'].sum()],
                '단가': ['']
            })
            result_df = pd.concat([result_df, total_row], ignore_index=True)
        
        return result_df
    
    def _create_full_inventory_list_sheet(self) -> pd.DataFrame:
        """전체재고리스트 시트 생성 (재고조사 후 계산값 적용)"""
        if self.inventory_data is None:
            return pd.DataFrame()
        
        # 실재고 데이터 복사
        result_df = self.inventory_data.copy()
        
        # 필요한 컬럼만 선택하고 컬럼명 정리
        if '제작사 품번' in result_df.columns:
            result_df = result_df[[
                '제작사 품번', '부품명', '단가', '재고', '재고액',
                '실재고', '실재고액', '차이', '차액'
            ]].copy()
            
            # 컬럼명 정리
            result_df.columns = [
                '제작사품번', '부품명', '단가', '재고', '재고액',
                '실재고', '실재고액', '차이', '차액'
            ]
        
        # 차액 절댓값 기준 내림차순 정렬 (차이가 큰 순서대로)
        result_df['차액_절댓값'] = result_df['차액'].abs()
        result_df = result_df.sort_values('차액_절댓값', ascending=False).reset_index(drop=True)
        result_df = result_df.drop('차액_절댓값', axis=1)  # 정렬용 컬럼 제거
        
        # 합계 행 추가
        if not result_df.empty:
            total_row = pd.DataFrame({
                '제작사품번': ['합계'],
                '부품명': [''],
                '단가': [''],
                '재고': [result_df['재고'].sum()],
                '재고액': [result_df['재고액'].sum()],
                '실재고': [result_df['실재고'].sum()],
                '실재고액': [result_df['실재고액'].sum()],
                '차이': [result_df['차이'].sum()],
                '차액': [result_df['차액'].sum()]
            })
            result_df = pd.concat([result_df, total_row], ignore_index=True)
        
        return result_df 