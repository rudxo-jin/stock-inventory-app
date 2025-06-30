import streamlit as st
import pandas as pd
import os
from datetime import datetime, date

# utils 모듈 import
try:
    from utils.data_processor import PartDataProcessor
    from utils.adjustment_processor import AdjustmentProcessor
    from utils.file_converter import ExcelFileConverter
    from utils.report_generator import ReportGenerator
except ImportError as e:
    st.error(f"모듈 import 오류: {e}")
    # 대안으로 직접 import 시도
    try:
        import sys
        import os
        sys.path.append(os.path.dirname(__file__))
        from utils.data_processor import PartDataProcessor
        from utils.adjustment_processor import AdjustmentProcessor
        from utils.file_converter import ExcelFileConverter
        from utils.report_generator import ReportGenerator
    except Exception as fallback_error:
        st.error(f"모듈 로드 실패: {fallback_error}")
        st.stop()

# 페이지 설정 (배포 최적화)
st.set_page_config(
    page_title="재고조사 앱",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 배포 환경 체크
@st.cache_data
def is_cloud_environment():
    """클라우드 배포 환경인지 확인"""
    return os.getenv('STREAMLIT_SHARING_MODE') is not None or os.getenv('DYNO') is not None

# 캐시된 프로세서 초기화 (배포 최적화)
@st.cache_resource
def get_processors(version="v2.1_error_fix"):
    """프로세서 인스턴스들을 캐시하여 메모리 효율성 향상"""
    try:
        return {
            'part_processor': PartDataProcessor(),
            'adjustment_processor': AdjustmentProcessor(),
            'report_generator': ReportGenerator()
        }
    except Exception as e:
        st.error(f"프로세서 초기화 오류: {str(e)}")
        st.error("앱을 새로고침해주세요.")
        st.stop()
        return None

# UI 컴포넌트 함수들 (UIComponents 대체)
def show_progress_sidebar():
    """사이드바에 진행 단계 표시"""
    with st.sidebar:
        st.markdown("### 📋 진행 단계")
        
        steps = [
            "1️⃣ PART 파일 업로드",
            "2️⃣ 실재고 템플릿 생성", 
            "3️⃣ 실재고 데이터 입력",
            "4️⃣ 재고조정 적용",
            "5️⃣ 결과보고서 생성"
        ]
        
        current_step = st.session_state.get('step', 1)
        
        for i, step_name in enumerate(steps, 1):
            if i <= current_step:
                st.success(step_name)
            elif i == current_step + 1:
                st.info(step_name)
            else:
                st.write(step_name)

def render_store_info_form():
    """점포 정보 입력 폼"""
    with st.form("store_info_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            store_name = st.text_input("점포명", value="고양점", placeholder="예: 고양점")
            survey_method = st.selectbox("조사방식", ["전수조사", "표본조사"], index=0)
        
        with col2:
            from datetime import date
            survey_date = st.date_input("재고조사일시", value=date.today())
            survey_staff = st.text_input("조사인원", value="", placeholder="조사에 참석한 직원명 표기")
        
        generate_report = st.form_submit_button("📋 보고서 생성", type="primary")
        
        if generate_report:
            return {
                'store_name': store_name,
                'survey_date': survey_date.strftime('%Y년 %m월 %d일'),
                'survey_method': survey_method,
                'survey_staff': survey_staff
            }
        return None

def get_card_styles():
    """카드 스타일 CSS를 반환"""
    return """
    <style>
    .card-container {
        display: flex;
        justify-content: center;
        margin: 0.5rem 0;
    }
    .section-card {
        background-color: #ffffff;
        padding: 1rem;
        border-radius: 12px;
        border: 1px solid #e0e0e0;
        margin: 0.3rem 0;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        max-width: 600px;
        width: 100%;
    }
    .section-title {
        font-size: 1.2rem;
        font-weight: bold;
        color: #333;
        margin-bottom: 0.8rem;
        padding-bottom: 0.3rem;
        border-bottom: 2px solid #f0f0f0;
        text-align: center;
    }
    .metric-table {
        width: 100%;
        border-collapse: collapse;
        max-width: 500px;
        margin: 0 auto;
    }
    .metric-row {
        border-bottom: 1px solid #f5f5f5;
    }
    .metric-row:last-child {
        border-bottom: none;
    }
    .metric-label {
        padding: 0.5rem 1rem;
        font-size: 1rem;
        color: #555;
        width: 50%;
        text-align: left;
    }
    .metric-value {
        padding: 0.5rem 1rem;
        font-size: 1.1rem;
        font-weight: bold;
        text-align: right;
        width: 50%;
    }
    .positive-value {
        color: #28a745;
    }
    .negative-value {
        color: #dc3545;
    }
    .total-value {
        color: #ffc107;
        background-color: #fffef8;
        padding: 0.6rem 1rem;
        border-radius: 6px;
    }
    .total-label {
        background-color: #fffef8;
        padding: 0.6rem 1rem;
        border-radius: 6px;
        font-weight: bold;
    }
    </style>
    """

def render_store_info_card(store_info):
    """점포 정보 카드를 렌더링"""
    return f"""
    <div class="card-container">
        <div class="section-card">
            <div class="section-title">🏪 점포 정보</div>
            <table class="metric-table">
                <tr class="metric-row">
                    <td class="metric-label">점포명</td>
                    <td class="metric-value">{store_info.get('store_name', '-')}</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label">조사일시</td>
                    <td class="metric-value">{store_info.get('survey_date', '-')}</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label">조사방식</td>
                    <td class="metric-value">{store_info.get('survey_method', '-')}</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label">조사인원</td>
                    <td class="metric-value">{store_info.get('survey_staff', '-')}</td>
                </tr>
            </table>
        </div>
    </div>
    """

def render_inventory_comparison_card(inv_comp):
    """전산재고 vs 실재고 카드를 렌더링"""
    return f"""
    <div class="card-container">
        <div class="section-card">
            <div class="section-title">📊 전산재고 vs 실재고</div>
            <table class="metric-table">
                <tr class="metric-row">
                    <td class="metric-label">전산재고액</td>
                    <td class="metric-value">{inv_comp.get('computer_stock_value', 0):,.0f}원</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label">(+) 실재고액</td>
                    <td class="metric-value positive-value">+{inv_comp.get('positive_amount', 0):,.0f}원</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label">(-) 실재고액</td>
                    <td class="metric-value negative-value">-{inv_comp.get('negative_amount', 0):,.0f}원</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label">최종재고액</td>
                    <td class="metric-value">{inv_comp.get('final_stock_value', 0):,.0f}원</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label total-label">차액</td>
                    <td class="metric-value total-value">{inv_comp.get('difference', 0):+,.0f}원</td>
                </tr>
            </table>
        </div>
    </div>
    """

def render_adjustment_impact_card(adj_imp):
    """재고조정 영향 카드를 렌더링"""
    return f"""
    <div class="card-container">
        <div class="section-card">
            <div class="section-title">⚖️ 재고조정 영향</div>
            <table class="metric-table">
                <tr class="metric-row">
                    <td class="metric-label">(+) 재고조정액</td>
                    <td class="metric-value positive-value">+{adj_imp.get('positive_adjustment', 0):,.0f}원</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label">(-) 재고조정액</td>
                    <td class="metric-value negative-value">-{adj_imp.get('negative_adjustment', 0):,.0f}원</td>
                </tr>
                <tr class="metric-row">
                    <td class="metric-label total-label">조정 차액</td>
                    <td class="metric-value total-value">{adj_imp.get('adjustment_difference', 0):+,.0f}원</td>
                </tr>
            </table>
        </div>
    </div>
    """

def render_total_impact_card(total_imp):
    """총 재고차액 카드를 렌더링 (안전한 렌더링)"""
    try:
        # 안전한 값 추출
        if not isinstance(total_imp, dict):
            total_imp = {}
            
        total_positive = total_imp.get('total_positive', 0)
        total_negative = total_imp.get('total_negative', 0)
        total_difference = total_imp.get('total_difference', 0)
        
        # 숫자 타입 확인
        if not isinstance(total_positive, (int, float)):
            total_positive = 0
        if not isinstance(total_negative, (int, float)):
            total_negative = 0
        if not isinstance(total_difference, (int, float)):
            total_difference = 0
        
        return f"""
        <div class="card-container">
            <div class="section-card">
                <div class="section-title">💰 총 재고차액</div>
                <table class="metric-table">
                    <tr class="metric-row">
                        <td class="metric-label">(+) 총재고차액</td>
                        <td class="metric-value positive-value">+{total_positive:,.0f}원</td>
                    </tr>
                    <tr class="metric-row">
                        <td class="metric-label">(-) 총재고차액</td>
                        <td class="metric-value negative-value">-{total_negative:,.0f}원</td>
                    </tr>
                    <tr class="metric-row">
                        <td class="metric-label total-label">총재고차액 계</td>
                        <td class="metric-value total-value">{total_difference:+,.0f}원</td>
                    </tr>
                </table>
            </div>
        </div>
        """
    except Exception as e:
        return f"""
        <div class="card-container">
            <div class="section-card">
                <div class="section-title">💰 총 재고차액</div>
                <div class="card-content">
                    <p style="color: red; text-align: center;">카드 렌더링 오류: {str(e)}</p>
                </div>
            </div>
        </div>
        """

def render_report_cards(report_data):
    """보고서 카드 렌더링 (원래 UIComponents와 동일한 스타일)"""
    try:
        # CSS 스타일 적용
        st.markdown(get_card_styles(), unsafe_allow_html=True)
        
        # 점포 정보 카드 (필수)
        if 'store_info' in report_data:
            st.markdown(render_store_info_card(report_data['store_info']), unsafe_allow_html=True)
        
        # 재고 비교 카드 (필수)
        if 'inventory_comparison' in report_data:
            st.markdown(render_inventory_comparison_card(report_data['inventory_comparison']), unsafe_allow_html=True)
        
        # 재고조정 영향 카드 (항상 표시 - 재고조정 적용 여부와 상관없이)
        adjustment_impact = report_data.get('adjustment_impact', {})
        # 빈 딕셔너리라도 카드는 표시 (0값으로)
        st.markdown(render_adjustment_impact_card(adjustment_impact), unsafe_allow_html=True)
        
        # 총 재고차액 카드 (항상 표시 - 안전한 렌더링)
        total_impact = report_data.get('total_impact', {})
        # 빈 딕셔너리라도 카드는 표시 (0값으로)
        st.markdown(render_total_impact_card(total_impact), unsafe_allow_html=True)
        

            
    except Exception as e:
        st.error(f"보고서 카드 렌더링 오류: {str(e)}")
        # 디버그 정보
        st.write("디버그 정보:")
        st.write(f"report_data 키들: {list(report_data.keys()) if isinstance(report_data, dict) else 'report_data가 딕셔너리가 아님'}")

def create_processed_inventory_excel(processed_data):
    """처리된 실재고 데이터를 엑셀 파일로 변환"""
    try:
        from io import BytesIO
        
        # 메모리에서 엑셀 파일 생성
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # 처리된 데이터를 엑셀로 저장
            processed_data.to_excel(writer, sheet_name='완성된실재고데이터', index=False)
            
            # 워크시트 스타일링
            workbook = writer.book
            worksheet = writer.sheets['완성된실재고데이터']
            
            # 스타일 임포트
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 색상 정의
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")    # 헤더용 진한 네이비
            positive_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # 양수용 연한 녹색
            negative_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")  # 음수용 연한 빨간색
            
            # 폰트 정의
            header_font = Font(bold=True, color="FFFFFF", size=11)  # 헤더용 흰색 볼드
            normal_font = Font(size=10)
            
            # 테두리 정의
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 헤더 스타일 적용
            for col_num in range(1, len(processed_data.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # 데이터 행 스타일 적용
            for row_num in range(2, len(processed_data) + 2):
                for col_num in range(1, len(processed_data.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 차이값 컬럼에 색상 적용 (차이 컬럼이 있을 경우)
                    if '차이' in processed_data.columns:
                        diff_col_idx = processed_data.columns.get_loc('차이') + 1
                        if col_num == diff_col_idx:
                            diff_value = processed_data.iloc[row_num - 2]['차이']
                            if diff_value > 0:
                                cell.fill = positive_fill
                            elif diff_value < 0:
                                cell.fill = negative_fill
                    
                    # 차액 컬럼에 색상 적용 (차액 컬럼이 있을 경우)
                    if '차액' in processed_data.columns:
                        amount_col_idx = processed_data.columns.get_loc('차액') + 1
                        if col_num == amount_col_idx:
                            amount_value = processed_data.iloc[row_num - 2]['차액']
                            if amount_value > 0:
                                cell.fill = positive_fill
                            elif amount_value < 0:
                                cell.fill = negative_fill
            
            # 컬럼 너비 자동 조정
            for col_num in range(1, len(processed_data.columns) + 1):
                column_letter = get_column_letter(col_num)
                col_name = processed_data.columns[col_num - 1]
                
                # 컬럼별 적절한 너비 설정
                if '제품명' in col_name:
                    worksheet.column_dimensions[column_letter].width = 30
                elif '제작사품번' in col_name:
                    worksheet.column_dimensions[column_letter].width = 20
                elif '차액' in col_name or '재고액' in col_name:
                    worksheet.column_dimensions[column_letter].width = 15
                else:
                    worksheet.column_dimensions[column_letter].width = 12
        
        # 바이트 데이터 반환
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"엑셀 파일 생성 오류: {str(e)}")
        return None

# 메인 함수
def main():
    st.title("📦 재고조사 앱")
    st.markdown("---")
    
    # 사이드바 진행 단계 표시
    show_progress_sidebar()
    
    # 캐시된 프로세서 가져오기
    processors = get_processors()
    
    # 프로세서 초기화 확인
    if processors is None:
        st.error("⚠️ 시스템 초기화 실패. 페이지를 새로고침해주세요.")
        st.stop()
        return
    
    # 세션 상태 초기화
    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'part_data' not in st.session_state:
        st.session_state.part_data = None
    if 'inventory_data' not in st.session_state:
        st.session_state.inventory_data = None
    if 'adjustment_data' not in st.session_state:
        st.session_state.adjustment_data = None
    if 'final_data' not in st.session_state:
        st.session_state.final_data = None
    if 'adjustment_summary' not in st.session_state:
        st.session_state.adjustment_summary = None
    # 엑셀 보고서 관련 세션 상태
    if 'excel_report_data' not in st.session_state:
        st.session_state.excel_report_data = None
    if 'excel_generation_time' not in st.session_state:
        st.session_state.excel_generation_time = None
    # 점포 정보 세션 상태
    if 'store_info' not in st.session_state:
        st.session_state.store_info = None
    
    # 탭 생성
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "1️⃣ PART 파일", 
        "2️⃣ 실재고 템플릿", 
        "3️⃣ 실재고 입력", 
        "4️⃣ 재고조정", 
        "5️⃣ 결과보고서"
    ])
    
    with tab1:
        st.header("📁 PART 파일 업로드")
        st.write("PART로 시작하는 엑셀 파일을 업로드해주세요.")
        
        uploaded_file = st.file_uploader(
            "PART 파일 선택",
            type=['xlsx', 'xls'],
            key="part_file"
        )
        
        if uploaded_file is not None:
            try:
                # 파일 자동 변환 처리
                converted_file_path = ExcelFileConverter.process_uploaded_file(uploaded_file)
                
                if converted_file_path:
                    st.success(f"✅ 파일 업로드 완료: {uploaded_file.name}")
                    
                    # 데이터 분석 버튼
                    if st.button("📊 데이터 분석하기", type="primary"):
                        with st.spinner("📊 PART 파일을 분석 중입니다..."):
                            success, message, data = processors['part_processor'].load_part_file(converted_file_path)
                            
                            if success:
                                st.session_state.part_data = data
                                st.session_state.step = 2
                                st.success(message)
                                
                                # 임시 파일 정리
                                ExcelFileConverter.cleanup_temp_file(converted_file_path)
                                st.rerun()
                            else:
                                st.error(message)
                                # 임시 파일 정리
                                ExcelFileConverter.cleanup_temp_file(converted_file_path)
                else:
                    st.error("❌ 파일 처리 실패")
                            
            except Exception as e:
                st.error(f"❌ 파일 업로드 오류: {str(e)}")
        
        # 분석 결과 표시
        if st.session_state.part_data is not None:
            st.markdown("### 📊 분석 결과")
            
            # 요약 통계 (안전한 접근)
            try:
                stats = processors['part_processor'].get_summary_stats()
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("총 품목 수", f"{stats.get('total_items', 0):,}개")
                    st.metric("재고 없는 품목", f"{stats.get('zero_stock_items', 0):,}개")
                with col2:
                    st.metric("총 재고량", f"{stats.get('total_stock', 0):,.0f}")
                    st.metric("평균 단가", f"{stats.get('avg_unit_price', 0):,.0f}원")
                with col3:
                    st.metric("총 재고액", f"{stats.get('total_stock_value', 0):,.0f}원")
                    st.metric("재고액 없는 품목", f"{stats.get('zero_value_items', 0):,}개")
            except Exception as e:
                st.error(f"통계 계산 오류: {str(e)}")
                # 기본값으로 표시
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("총 품목 수", "0개")
                    st.metric("재고 없는 품목", "0개")
                with col2:
                    st.metric("총 재고량", "0")
                    st.metric("평균 단가", "0원")
                with col3:
                    st.metric("총 재고액", "0원")
                    st.metric("재고액 없는 품목", "0개")
            
            # 데이터 미리보기
            st.markdown("### 📋 데이터 미리보기")
            st.dataframe(
                st.session_state.part_data.head(10),
                use_container_width=True
            )
    
    with tab2:
        st.header("📋 실재고 입력 템플릿")
        if st.session_state.step >= 2 and st.session_state.part_data is not None:
            st.write("PART 파일 분석이 완료되었습니다. 실재고 입력용 템플릿을 다운로드하세요.")
            
            # 템플릿 생성 및 다운로드
            if st.button("📥 템플릿 생성", type="primary"):
                try:
                    template = processors['part_processor'].create_inventory_template()
                    
                    # 메모리에서 엑셀 파일 생성 (웹 배포 호환)
                    from io import BytesIO
                    buffer = BytesIO()
                    template.to_excel(buffer, index=False, engine='openpyxl')
                    buffer.seek(0)
                    
                    st.success("✅ 템플릿이 생성되었습니다!")
                    
                    # 파일 다운로드 버튼
                    st.download_button(
                        label="📥 템플릿 다운로드",
                        data=buffer.getvalue(),
                        file_name="실재고입력템플릿.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # 템플릿 미리보기
                    st.markdown("### 📋 템플릿 미리보기")
                    st.dataframe(template.head(10), use_container_width=True)
                    
                    # 템플릿 정보 표시
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("템플릿 품목 수", f"{len(template):,}개")
                    with col2:
                        excluded_count = len(st.session_state.part_data) - len(template)
                        st.metric("제외된 품목 수", f"{excluded_count:,}개 (재고 없음)")
                    
                    # 사용 안내
                    st.markdown("### 📝 사용 안내")
                    st.info("""
                    **실재고 입력 방법:**
                    - **실재고** 컬럼에 실제 재고량을 입력하거나
                    - **차이** 컬럼에 차이값을 입력하세요 (예: -5, +10)
                    - 둘 다 입력된 경우 **차이값이 우선**됩니다
                    - 입력하지 않은 품목은 기존 재고로 유지됩니다
                    - **재고가 없는 품목은 템플릿에서 제외**되었습니다
                    """)
                    
                    st.session_state.step = 3
                    
                except Exception as e:
                    st.error(f"❌ 템플릿 생성 오류: {str(e)}")
        else:
            st.warning("⚠️ 먼저 PART 파일을 업로드하고 분석해주세요.")
    
    with tab3:
        st.header("📝 실재고 데이터 입력")
        if st.session_state.step >= 3:
            st.write("작성된 실재고 데이터를 업로드해주세요.")
            
            uploaded_inventory = st.file_uploader(
                "실재고 파일 선택",
                type=['xlsx', 'xls'],
                key="inventory_file"
            )
            
            if uploaded_inventory is not None:
                try:
                    # 파일 자동 변환 처리
                    converted_file_path = ExcelFileConverter.process_uploaded_file(uploaded_inventory)
                    
                    if converted_file_path:
                        # 파일 읽기
                        inventory_df = pd.read_excel(converted_file_path, engine='openpyxl')
                        
                        # 데이터 검증 및 계산
                        success, message, processed_data = processors['part_processor'].validate_inventory_data(inventory_df)
                        
                        if success:
                            st.session_state.inventory_data = processed_data
                            st.session_state.step = 4
                            st.success(message)
                            
                            # 처리 결과 미리보기
                            st.markdown("### 📊 처리 결과")
                            
                            # 요약 통계
                            total_items = len(processed_data)
                            changed_items = len(processed_data[processed_data['차이'] != 0])
                            positive_diff = processed_data[processed_data['차이'] > 0]['차이'].sum()
                            negative_diff = processed_data[processed_data['차이'] < 0]['차이'].sum()
                            total_diff_value = processed_data['차액'].sum()
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("총 품목 수", f"{total_items:,}개")
                                st.metric("변경된 품목", f"{changed_items:,}개")
                            with col2:
                                st.metric("증가 수량", f"{positive_diff:,.0f}")
                                st.metric("감소 수량", f"{negative_diff:,.0f}")
                            with col3:
                                st.metric("총 차액", f"{total_diff_value:,.0f}원")
                            
                            # 데이터 미리보기
                            st.markdown("### 📋 데이터 미리보기")
                            st.dataframe(processed_data.head(10), use_container_width=True)
                            
                            # 처리된 데이터 다운로드 기능 추가
                            st.markdown("### 📥 완성된 실재고 파일 다운로드")
                            st.write("계산이 완료된 실재고 데이터를 엑셀 파일로 다운로드할 수 있습니다.")
                            
                            # 엑셀 파일 생성
                            excel_data = create_processed_inventory_excel(processed_data)
                            if excel_data:
                                st.download_button(
                                    label="📊 완성된 실재고 파일 다운로드",
                                    data=excel_data,
                                    file_name=f"완성된_실재고데이터_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary",
                                    help="처리가 완료된 실재고 데이터를 엑셀 파일로 다운로드합니다."
                                )
                                st.success("✅ 엑셀 파일이 준비되었습니다. 위 버튼을 클릭하여 다운로드하세요.")
                            else:
                                st.error("❌ 엑셀 파일 생성에 실패했습니다.")
                            
                        else:
                            st.error(message)
                        
                        # 임시 파일 정리
                        ExcelFileConverter.cleanup_temp_file(converted_file_path)
                    else:
                        st.error("❌ 파일 처리 실패")
                        
                except Exception as e:
                    st.error(f"❌ 파일 처리 오류: {str(e)}")
        else:
            st.warning("⚠️ 먼저 이전 단계를 완료해주세요.")
    
    with tab4:
        st.header("⚖️ 재고조정 적용")
        st.write("재고조정 파일을 업로드하고 기간을 설정하여 적용할 수 있습니다. (선택사항)")
        
        # 재고조정 파일 업로드
        uploaded_adjustment = st.file_uploader(
            "재고조정 파일 선택",
            type=['xlsx', 'xls'],
            key="adjustment_file"
        )
        
        if uploaded_adjustment is not None:
            try:
                # 파일 자동 변환 처리
                converted_file_path = ExcelFileConverter.process_uploaded_file(uploaded_adjustment)
                
                if converted_file_path:
                    st.success(f"✅ 파일 업로드 완료: {uploaded_adjustment.name}")
                    
                    # 재고조정 파일 로드
                    success, message, adj_data = processors['adjustment_processor'].load_adjustment_file(converted_file_path)
                    
                    if success:
                        st.session_state.adjustment_data = adj_data
                        st.success(message)
                        
                        # 데이터 미리보기
                        st.markdown("### 📋 재고조정 데이터 미리보기")
                        st.dataframe(adj_data.head(10), use_container_width=True)
                        
                        # 날짜 범위 설정
                        st.markdown("### 📅 적용 기간 설정")
                        
                        # 폼으로 날짜 입력과 적용 버튼을 함께 처리
                        with st.form("adjustment_form"):
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                # 기본값을 조정 데이터의 최소/최대 날짜로 설정
                                min_date = adj_data['일자'].min().date() if not adj_data.empty else date.today()
                                max_date = adj_data['일자'].max().date() if not adj_data.empty else date.today()
                                
                                start_date = st.date_input(
                                    "시작일",
                                    value=min_date,
                                    min_value=date(1900, 1, 1),
                                    max_value=date(2100, 12, 31)
                                )
                            
                            with col2:
                                end_date = st.date_input(
                                    "종료일",
                                    value=max_date,
                                    min_value=date(1900, 1, 1),
                                    max_value=date(2100, 12, 31)
                                )
                            
                            # 재고조정 적용 버튼
                            apply_adjustment = st.form_submit_button("⚖️ 재고조정 적용", type="primary")
                        
                        if apply_adjustment:
                            if start_date <= end_date:
                                with st.spinner("⚖️ 재고조정을 적용 중입니다..."):
                                    # 날짜 범위로 필터링
                                    filter_success, filter_message, filtered_data = processors['adjustment_processor'].filter_by_date_range(start_date, end_date)
                                    
                                    if filter_success and st.session_state.inventory_data is not None:
                                        # 실재고 데이터에 재고조정 적용
                                        apply_success, apply_message, final_data, adj_summary = processors['adjustment_processor'].apply_adjustments_to_inventory(
                                            st.session_state.inventory_data, st.session_state.part_data
                                        )
                                        
                                        if apply_success:
                                            st.session_state.final_data = final_data
                                            st.session_state.adjustment_summary = adj_summary
                                            st.session_state.step = max(st.session_state.step, 4)
                                            st.success(apply_message)
                                            
                                            # 적용 결과 표시
                                            st.markdown("### 📊 재고조정 적용 결과")
                                            
                                            col1, col2, col3, col4 = st.columns(4)
                                            with col1:
                                                st.metric("필터된 조정 건수", f"{len(filtered_data):,}건")
                                            with col2:
                                                st.metric("매칭된 조정 건수", f"{adj_summary.get('total_adjustments', 0):,}건")
                                            with col3:
                                                positive_amt = adj_summary.get('positive_amount', 0)
                                                st.metric("(+) 조정액", f"{positive_amt:,.0f}원")
                                            with col4:
                                                negative_amt = abs(adj_summary.get('negative_amount', 0))
                                                st.metric("(-) 조정액", f"{negative_amt:,.0f}원")
                                            
                                            # 미매칭 품목이 있으면 경고 표시
                                            unmatched_items = adj_summary.get('unmatched_items', [])
                                            if unmatched_items:
                                                st.warning(f"⚠️ {len(unmatched_items)}개 품목이 실재고 데이터와 매칭되지 않았습니다.")
                                                
                                                # 미매칭 품목 상세 정보 (접기/펼치기)
                                                with st.expander("미매칭 품목 상세"):
                                                    unmatched_df = pd.DataFrame(unmatched_items)
                                                    st.dataframe(unmatched_df, use_container_width=True)
                                            
                                            st.rerun()
                                        else:
                                            st.error(apply_message)
                                    elif not filter_success:
                                        st.error(filter_message)
                                    else:
                                        st.warning("⚠️ 먼저 실재고 데이터를 입력해주세요.")
                            else:
                                st.error("❌ 시작일이 종료일보다 늦을 수 없습니다.")
                    else:
                        st.error(message)
                    
                    # 임시 파일 정리
                    ExcelFileConverter.cleanup_temp_file(converted_file_path)
                else:
                    st.error("❌ 파일 처리 실패")
                    
            except Exception as e:
                st.error(f"❌ 파일 처리 오류: {str(e)}")
        
        # 재고조정 요약 정보 표시 (적용된 경우에만)
        if st.session_state.adjustment_summary is not None:
            st.markdown("### 📋 재고조정 요약")
            summary = processors['adjustment_processor'].get_adjustment_summary()
            
            if summary:
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("(+) 조정 건수", f"{summary.get('positive_records', 0):,}건")
                    st.metric("(+) 조정 수량", f"{summary.get('positive_quantity', 0):,.0f}")
                with col2:
                    st.metric("(-) 조정 건수", f"{summary.get('negative_records', 0):,}건")
                    st.metric("(-) 조정 수량", f"{abs(summary.get('negative_quantity', 0)):,.0f}")
    
    with tab5:
        st.header("📊 결과보고서")
        
        # 조건 확인 (step >= 3이고 inventory_data가 있으면 OK)
        if st.session_state.step >= 3 and st.session_state.inventory_data is not None:
            # 점포 정보 입력 및 세션 저장
            store_info = render_store_info_form()
            
            # 점포 정보가 입력되면 세션에 저장
            if store_info:
                st.session_state.store_info = store_info
            
            # 세션에 저장된 점포 정보 사용 (폼 제출과 무관하게 유지)
            if hasattr(st.session_state, 'store_info') and st.session_state.store_info:
                # final_data가 없으면 inventory_data 사용
                report_data_source = st.session_state.final_data if st.session_state.final_data is not None else st.session_state.inventory_data
                
                # 재고조정 데이터 설정
                filtered_adj_data = getattr(processors['adjustment_processor'], 'filtered_data', None)
                if filtered_adj_data is not None:
                    processors['report_generator'].set_adjustment_data(filtered_adj_data)
                elif st.session_state.adjustment_data is not None:
                    processors['report_generator'].set_adjustment_data(st.session_state.adjustment_data)
                
                # 보고서 데이터 생성 (세션에 저장된 점포 정보 사용)
                report_data = processors['report_generator'].generate_report_data(
                    inventory_data=report_data_source,
                    store_info=st.session_state.store_info,
                    part_data=st.session_state.part_data,
                    final_data=st.session_state.final_data,
                    adjustment_summary=st.session_state.adjustment_summary
                )
                
                if report_data:
                    # 보고서 카드 표시
                    render_report_cards(report_data)
                    
                    # 요약 통계
                    stats = processors['report_generator'].get_summary_stats()
                    
                    # 엑셀 보고서 다운로드
                    st.markdown("### 📥 보고서 다운로드")
                    
                    # 엑셀 보고서 생성 (세션 상태 기반)
                    col1, col2 = st.columns([1, 3])
                    with col1:
                        # 엑셀 생성 버튼 (완전 단순화)
                        if st.button("📊 엑셀 보고서 생성", type="primary", key="generate_excel"):
                            try:
                                with st.spinner("📊 엑셀 보고서를 생성 중입니다..."):
                                    # 현재 화면에 표시된 report_data 사용 (재생성 안함)
                                    report_generator = processors['report_generator']
                                    
                                    # 엑셀 보고서 생성 (별도 검증 없이 바로 시도)
                                    excel_data = report_generator.create_excel_report()
                                    
                                    if excel_data and len(excel_data) > 0:
                                        # 세션에 저장
                                        st.session_state.excel_report_data = excel_data
                                        st.session_state.excel_generation_time = datetime.now().strftime("%Y%m%d_%H%M%S")
                                        st.success("✅ 엑셀 보고서가 성공적으로 생성되었습니다!")
                                        
                                        # 즉시 다운로드 버튼 표시
                                        filename = f"재고조사보고서_{st.session_state.excel_generation_time}.xlsx"
                                        
                                        st.download_button(
                                            label="📥 보고서 다운로드",
                                            data=st.session_state.excel_report_data,
                                            file_name=filename,
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            key="download_excel_immediate"
                                        )
                                        
                                        # 파일 크기 정보
                                        file_size = len(st.session_state.excel_report_data) / 1024  # KB
                                        st.info(f"📄 파일 크기: {file_size:.1f}KB")
                                    else:
                                        st.error("❌ 엑셀 보고서 생성 실패: 데이터가 비어있습니다.")
                                        
                            except Exception as e:
                                st.error(f"❌ 보고서 생성 오류: {str(e)}")
                                st.error("점포 정보를 다시 입력하고 보고서를 먼저 생성해주세요.")
                        
                        # 기존 다운로드 버튼 (엑셀 데이터가 있을 때만 표시)
                        elif st.session_state.excel_report_data is not None:
                            filename = f"재고조사보고서_{st.session_state.excel_generation_time}.xlsx"
                            
                            st.download_button(
                                label="📥 보고서 다운로드 (기존)",
                                data=st.session_state.excel_report_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_excel_persistent"
                            )
                            
                            # 파일 크기 정보
                            file_size = len(st.session_state.excel_report_data) / 1024  # KB
                            st.info(f"📄 파일 크기: {file_size:.1f}KB")
                            
                            # 새 보고서 생성 버튼
                            if st.button("🔄 새 보고서 생성", key="reset_excel"):
                                st.session_state.excel_report_data = None
                                st.session_state.excel_generation_time = None
                    
                    with col2:
                        st.info("📋 **보고서 구성**: 요약보고서, 재고차이리스트, 재고조정리스트 (5개 시트)")
                        
                        # 현재 상태 표시
                        if st.session_state.excel_report_data is not None:
                            st.success(f"✅ 엑셀 파일 준비 완료 ({st.session_state.excel_generation_time})")
                        else:
                            st.info("💡 먼저 '엑셀 보고서 생성' 버튼을 클릭하세요")
                        
                        # 엑셀 보고서 설명
                        with st.expander("📝 엑셀 보고서 상세 내용"):
                            st.write("""
                            **포함 시트:**
                            - 📊 재고조사요약: 전체 결과 요약
                            - 📉 재고차이리스트(-): 부족 재고 상세
                            - 📈 재고차이리스트(+): 과잉 재고 상세
                            - ⚖️ 재고조정리스트(+): 증가 조정 내역
                            - ⚖️ 재고조정리스트(-): 감소 조정 내역
                            """)
                else:
                    st.error("❌ 보고서 데이터 생성 실패")
            else:
                st.info("💡 위의 점포 정보를 입력하고 '📋 보고서 생성' 버튼을 클릭해주세요.")
        else:
            st.warning("⚠️ 먼저 이전 단계를 완료해주세요.")
            
            # 디버깅 정보 (개발용)
            if st.checkbox("🔧 디버깅 정보 표시"):
                st.write(f"현재 단계: {st.session_state.step}")
                st.write(f"PART 데이터: {'있음' if st.session_state.part_data is not None else '없음'}")
                st.write(f"실재고 데이터: {'있음' if st.session_state.inventory_data is not None else '없음'}")
                st.write(f"최종 데이터: {'있음' if st.session_state.final_data is not None else '없음'}")
                
                # 강제 보고서 생성 (개발/테스트용)
                if st.session_state.inventory_data is not None:
                    if st.button("🔧 강제 보고서 생성 (테스트용)"):
                        st.session_state.step = 5
                        st.rerun()

if __name__ == "__main__":
    main() 